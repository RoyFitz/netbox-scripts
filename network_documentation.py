"""
NetBox Custom Script: Network Documentation Generator

Generates an Excel document with network documentation for a selected site.
Includes cover page, summary of prefixes/VLANs, and per-prefix device listings.

Place this file in: /opt/netbox/netbox/scripts/
"""

from extras.scripts import Script, ObjectVar, BooleanVar
from dcim.models import Site
from ipam.models import Prefix, VLAN, IPAddress
from dcim.models import Interface
from virtualization.models import VMInterface
from django.db.models import Q
from datetime import datetime
from io import BytesIO

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Check for NetBox Branching plugin
try:
    from netbox_branching.models import Branch
    from netbox_branching.utilities import activate_branch
    BRANCHING_AVAILABLE = True
except ImportError:
    BRANCHING_AVAILABLE = False
    Branch = None


class NetworkDocumentationScript(Script):
    """Generate Excel network documentation for a site."""

    class Meta:
        name = "Network Documentation Generator"
        description = "Generates an Excel document with network documentation for a selected site"
        job_timeout = 300  # 5 minutes max

    # ==========================================================================
    # Script Variables (Form Fields)
    # ==========================================================================

    site = ObjectVar(
        model=Site,
        required=True,
        description="Select the site to generate documentation for"
    )

    include_empty_prefixes = BooleanVar(
        default=True,
        description="Include prefixes with no IP addresses assigned"
    )

    show_unused_ips = BooleanVar(
        default=True,
        description="Show unused/available IP addresses in each prefix (only for prefixes /24 or smaller)"
    )

    # Branch selector (only shown if netbox_branching plugin is available)
    if BRANCHING_AVAILABLE:
        branch = ObjectVar(
            model=Branch,
            required=False,
            description="Select a branch to query data from (leave empty for main/production)"
        )

    # ==========================================================================
    # Styles
    # ==========================================================================

    def _init_styles(self):
        """Initialize Excel styles for consistent formatting."""
        self.log_debug("Initializing Excel styles")

        # Colors
        self.HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        self.ALT_ROW_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
        self.ORPHAN_HEADER_FILL = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
        self.GATEWAY_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # Gold/yellow
        self.GATEWAY_FONT = Font(name="Calibri", size=11, bold=True)

        # Fonts
        self.TITLE_FONT = Font(name="Calibri", size=28, bold=True, color="1F4E79")
        self.SUBTITLE_FONT = Font(name="Calibri", size=14, color="666666")
        self.HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        self.NORMAL_FONT = Font(name="Calibri", size=11)
        self.SECTION_FONT = Font(name="Calibri", size=14, bold=True, color="1F4E79")

        # Borders
        thin_side = Side(style='thin', color='B4B4B4')
        self.CELL_BORDER = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        # Alignment
        self.CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
        self.LEFT_ALIGN = Alignment(horizontal='left', vertical='center')

    # ==========================================================================
    # Data Collection Methods
    # ==========================================================================

    def _get_site_prefixes(self, site):
        """Retrieve all prefixes for the site."""
        self.log_debug(f"Querying prefixes for site: {site.name}")

        prefixes = Prefix.objects.filter(site=site).select_related('vlan', 'role').order_by('prefix')
        prefix_count = prefixes.count()

        self.log_info(f"Found {prefix_count} prefixes for site '{site.name}'")

        if prefix_count == 0:
            self.log_warning(f"No prefixes found for site '{site.name}' - check if prefixes have site assigned")

        return prefixes

    def _get_site_vlans(self, site):
        """Retrieve all VLANs for the site."""
        self.log_debug(f"Querying VLANs for site: {site.name}")

        vlans = VLAN.objects.filter(site=site).select_related('role').order_by('vid')
        vlan_count = vlans.count()

        self.log_info(f"Found {vlan_count} VLANs for site '{site.name}'")

        if vlan_count == 0:
            self.log_warning(f"No VLANs found for site '{site.name}' - check if VLANs have site assigned")

        return vlans

    def _get_orphan_vlans(self, site, prefixes):
        """Find VLANs not associated with any prefix."""
        self.log_debug("Identifying orphan VLANs (not associated with any prefix)")

        # Get all VLAN IDs that are associated with prefixes
        prefix_vlan_ids = set(
            prefixes.exclude(vlan__isnull=True).values_list('vlan_id', flat=True)
        )
        self.log_debug(f"VLANs associated with prefixes: {prefix_vlan_ids}")

        # Get all site VLANs not in that set
        orphan_vlans = VLAN.objects.filter(site=site).exclude(id__in=prefix_vlan_ids).order_by('vid')
        orphan_count = orphan_vlans.count()

        if orphan_count > 0:
            self.log_warning(f"Found {orphan_count} orphan VLANs (not associated with any prefix)")
            for vlan in orphan_vlans:
                self.log_debug(f"  Orphan VLAN: {vlan.vid} - {vlan.name}")
        else:
            self.log_info("No orphan VLANs found - all VLANs are associated with prefixes")

        return orphan_vlans

    def _is_default_gateway(self, ip_address):
        """Check if an IP address has the 'default-gateway' tag."""
        try:
            return ip_address.tags.filter(slug='default-gateway').exists()
        except Exception:
            return False

    def _get_prefix_ip_addresses(self, prefix):
        """Get all IP addresses within a prefix with their assigned devices."""
        import netaddr

        prefix_network = netaddr.IPNetwork(str(prefix.prefix))

        # Get all IPs and filter manually (works reliably with netbox_branching)
        all_ips = list(IPAddress.objects.all())

        # Filter IPs that fall within this prefix
        matching_ips = []
        for ip in all_ips:
            try:
                ip_addr = netaddr.IPAddress(str(ip.address).split('/')[0])
                if ip_addr in prefix_network:
                    matching_ips.append(ip)
            except Exception:
                pass

        # Sort by address
        matching_ips.sort(key=lambda x: netaddr.IPAddress(str(x.address).split('/')[0]))

        self.log_debug(f"Prefix {prefix.prefix}: found {len(matching_ips)} IPs")

        return matching_ips

    def _get_ip_device_info(self, ip_address):
        """
        Extract device/VM information from an IP address assignment.
        Returns dict with device_name, device_role, device_model, interface_name, device_type.
        """
        result = {
            'device_name': '',
            'device_role': '',
            'device_model': '',
            'interface_name': '',
            'device_type': '',
            'status': str(ip_address.status) if ip_address.status else ''
        }

        assigned_object = ip_address.assigned_object

        if assigned_object is None:
            return result

        try:
            if isinstance(assigned_object, Interface):
                # Physical device interface
                device = assigned_object.device
                result['device_name'] = device.name if device else ''
                result['device_role'] = device.role.name if device and device.role else ''
                result['device_model'] = device.device_type.model if device and device.device_type else ''
                result['interface_name'] = assigned_object.name
                result['device_type'] = 'Device'

            elif isinstance(assigned_object, VMInterface):
                # Virtual machine interface
                vm = assigned_object.virtual_machine
                result['device_name'] = vm.name if vm else ''
                result['device_role'] = vm.role.name if vm and vm.role else ''
                result['device_model'] = vm.platform.name if vm and vm.platform else 'Virtual'
                result['interface_name'] = assigned_object.name
                result['device_type'] = 'VM'

            else:
                # Some other assignment type
                result['device_name'] = str(assigned_object)
                result['device_type'] = type(assigned_object).__name__
                self.log_debug(f"IP {ip_address.address} -> Other: {result['device_type']}")

        except Exception as e:
            self.log_warning(f"Error getting device info for IP {ip_address.address}: {str(e)}")

        return result

    # ==========================================================================
    # Excel Sheet Creation Methods
    # ==========================================================================

    def _create_cover_page(self, workbook, site):
        """Create the cover page worksheet."""
        self.log_info("Creating cover page")

        ws = workbook.active
        ws.title = "Cover"

        # Set column width
        ws.column_dimensions['A'].width = 60

        # Title
        ws['A5'] = "Network Documentation"
        ws['A5'].font = self.TITLE_FONT
        ws['A5'].alignment = self.CENTER_ALIGN

        # Site name
        ws['A8'] = site.name
        ws['A8'].font = Font(name="Calibri", size=22, bold=True)
        ws['A8'].alignment = self.CENTER_ALIGN

        # Site details
        row = 11
        details = [
            ("Site Status", str(site.status) if site.status else "N/A"),
            ("Region", site.region.name if site.region else "N/A"),
            ("Facility", site.facility if site.facility else "N/A"),
            ("Physical Address", site.physical_address if site.physical_address else "N/A"),
            ("ASN", str(site.asns.first()) if site.asns.exists() else "N/A"),
        ]

        for label, value in details:
            ws[f'A{row}'] = f"{label}:"
            ws[f'A{row}'].font = Font(name="Calibri", size=12, bold=True)
            ws[f'A{row + 1}'] = value
            ws[f'A{row + 1}'].font = self.SUBTITLE_FONT
            ws[f'A{row + 1}'].alignment = Alignment(wrap_text=True)
            row += 3

        # Generation timestamp
        ws[f'A{row + 2}'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws[f'A{row + 2}'].font = self.SUBTITLE_FONT
        ws[f'A{row + 2}'].alignment = self.CENTER_ALIGN

        self.log_debug("Cover page created successfully")

    def _create_summary_sheet(self, workbook, site, prefixes, orphan_vlans, prefix_sheet_names):
        """Create the summary worksheet with prefix/VLAN overview."""
        self.log_info("Creating summary sheet")

        ws = workbook.create_sheet("Summary")

        # Set column widths
        col_widths = [12, 25, 20, 18, 40, 18]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Title
        ws['A1'] = f"Network Summary - {site.name}"
        ws['A1'].font = self.SECTION_FONT
        ws.merge_cells('A1:F1')

        # Prefixes section header
        current_row = 3
        ws[f'A{current_row}'] = "Prefixes and Associated VLANs"
        ws[f'A{current_row}'].font = self.SECTION_FONT

        # Table headers
        current_row += 1
        headers = ["VLAN ID", "VLAN Name", "Prefix", "Gateway", "Description", "Used/Available"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.CENTER_ALIGN
            cell.border = self.CELL_BORDER

        # Prefix data rows
        current_row += 1
        prefix_start_row = current_row
        prefixes_with_data = 0

        # Link font style (blue, underlined)
        link_font = Font(name="Calibri", size=11, color="0563C1", underline="single")

        # Track utilization data for data bars
        utilization_percentages = []

        for prefix in prefixes:
            try:
                # Calculate utilization as used/total count
                try:
                    import netaddr
                    prefix_network = netaddr.IPNetwork(str(prefix.prefix))
                    # Total usable IPs (excluding network and broadcast for /30 or larger)
                    if prefix_network.prefixlen <= 30:
                        total_ips = prefix_network.size - 2
                    else:
                        total_ips = prefix_network.size  # /31 and /32 use all addresses

                    # Count assigned IPs using our method
                    assigned_ips = self._get_prefix_ip_addresses(prefix)
                    used_ips = len(assigned_ips)

                    utilization = f"{used_ips} / {total_ips}"
                    utilization_pct = (used_ips / total_ips * 100) if total_ips > 0 else 0

                    # Find default gateway IP
                    gateway_ips = [ip for ip in assigned_ips if self._is_default_gateway(ip)]
                    gateway_str = str(gateway_ips[0].address).split('/')[0] if gateway_ips else ""
                except Exception:
                    utilization = "N/A"
                    utilization_pct = 0
                    gateway_str = ""

                row_data = [
                    prefix.vlan.vid if prefix.vlan else "None",
                    prefix.vlan.name if prefix.vlan else "No VLAN",
                    str(prefix.prefix),
                    gateway_str,
                    prefix.description or "",
                    utilization_pct  # Percentage for data bar
                ]

                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col, value=value)
                    cell.font = self.NORMAL_FONT
                    cell.border = self.CELL_BORDER
                    cell.alignment = self.LEFT_ALIGN
                    # Format the Used/Available column to show count with data bar
                    if col == 6:
                        cell.number_format = f'"{utilization}"'
                        cell.alignment = self.CENTER_ALIGN

                # Store row for data bar range
                utilization_percentages.append(current_row)

                # Make prefix cell a hyperlink to its worksheet
                sheet_name = prefix_sheet_names.get(prefix.id)
                if sheet_name:
                    from openpyxl.worksheet.hyperlink import Hyperlink
                    prefix_cell = ws.cell(row=current_row, column=3)
                    prefix_cell.hyperlink = Hyperlink(ref=prefix_cell.coordinate, location=f"'{sheet_name}'!A1")
                    prefix_cell.font = link_font

                # Alternate row coloring
                if (current_row - prefix_start_row) % 2 == 1:
                    for col in range(1, 7):
                        cell = ws.cell(row=current_row, column=col)
                        cell.fill = self.ALT_ROW_FILL
                        # Preserve link styling for prefix column
                        if col == 3 and sheet_name:
                            cell.font = link_font

                current_row += 1
                prefixes_with_data += 1

            except Exception as e:
                self.log_warning(f"Error processing prefix {prefix.prefix}: {str(e)}")

        # Add data bar conditional formatting to the Used/Available column (column 6)
        if utilization_percentages:
            from openpyxl.formatting.rule import DataBarRule

            first_row = utilization_percentages[0]
            last_row = utilization_percentages[-1]

            # Create data bar rule (green bar, 0-100 scale)
            data_bar_rule = DataBarRule(
                start_type='num',
                start_value=0,
                end_type='num',
                end_value=100,
                color='63BE7B',  # Green color
                minLength=0  # No bar when value is 0
            )

            ws.conditional_formatting.add(f'F{first_row}:F{last_row}', data_bar_rule)

        self.log_info(f"Added {prefixes_with_data} prefixes to summary")

        # Orphan VLANs section
        if orphan_vlans.exists():
            current_row += 2
            ws[f'A{current_row}'] = "Orphan VLANs (Not Associated with Prefixes)"
            ws[f'A{current_row}'].font = self.SECTION_FONT

            current_row += 1
            orphan_headers = ["VLAN ID", "VLAN Name", "Description", "Status"]
            for col, header in enumerate(orphan_headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = self.HEADER_FONT
                cell.fill = self.ORPHAN_HEADER_FILL
                cell.alignment = self.CENTER_ALIGN
                cell.border = self.CELL_BORDER

            current_row += 1
            orphan_start_row = current_row

            for vlan in orphan_vlans:
                try:
                    row_data = [
                        vlan.vid,
                        vlan.name,
                        vlan.description or "",
                        str(vlan.status) if vlan.status else ""
                    ]

                    for col, value in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col, value=value)
                        cell.font = self.NORMAL_FONT
                        cell.border = self.CELL_BORDER

                    if (current_row - orphan_start_row) % 2 == 1:
                        for col in range(1, len(orphan_headers) + 1):
                            ws.cell(row=current_row, column=col).fill = self.ALT_ROW_FILL

                    current_row += 1

                except Exception as e:
                    self.log_warning(f"Error processing orphan VLAN {vlan.vid}: {str(e)}")

            self.log_info(f"Added {orphan_vlans.count()} orphan VLANs to summary")

        self.log_debug("Summary sheet created successfully")

    def _create_prefix_sheets(self, workbook, prefixes, include_empty, prefix_sheet_names, show_unused_ips):
        """Create individual worksheets for each prefix."""
        import netaddr

        self.log_info(f"Creating prefix detail sheets (include_empty={include_empty}, show_unused={show_unused_ips})")

        # Style for available/unused IPs
        available_font = Font(name="Calibri", size=11, italic=True, color="808080")
        available_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

        sheets_created = 0
        sheets_skipped = 0

        for prefix in prefixes:
            try:
                # Get assigned IP addresses for this prefix
                assigned_ips = self._get_prefix_ip_addresses(prefix)

                # Skip empty prefixes if configured (and not showing unused)
                if not include_empty and len(assigned_ips) == 0 and not show_unused_ips:
                    self.log_debug(f"Skipping empty prefix: {prefix.prefix}")
                    sheets_skipped += 1
                    continue

                # Get pre-calculated sheet name
                sheet_name = prefix_sheet_names.get(prefix.id, str(prefix.prefix).replace('/', '_')[:31])
                self.log_debug(f"Creating sheet for prefix: {prefix.prefix} as '{sheet_name}'")

                ws = workbook.create_sheet(sheet_name)

                # Set column widths
                col_widths = [18, 25, 18, 22, 18, 10, 12]
                for i, width in enumerate(col_widths, 1):
                    ws.column_dimensions[get_column_letter(i)].width = width

                # Find default gateway IP(s) in this prefix
                gateway_ips = [ip for ip in assigned_ips if self._is_default_gateway(ip)]
                gateway_str = ", ".join(str(ip.address).split('/')[0] for ip in gateway_ips) if gateway_ips else "N/A"

                # Prefix header info
                ws['A1'] = f"Prefix: {prefix.prefix}"
                ws['A1'].font = self.SECTION_FONT

                ws['A2'] = f"VLAN: {prefix.vlan.vid} - {prefix.vlan.name}" if prefix.vlan else "VLAN: None"
                ws['A2'].font = self.SUBTITLE_FONT

                ws['A3'] = f"Description: {prefix.description or 'N/A'}"
                ws['A3'].font = self.SUBTITLE_FONT

                ws['A4'] = f"Role: {prefix.role.name if prefix.role else 'N/A'}"
                ws['A4'].font = self.SUBTITLE_FONT

                # Default Gateway row (highlighted)
                ws['A5'] = f"Default Gateway: {gateway_str}"
                ws['A5'].font = self.GATEWAY_FONT
                if gateway_ips:
                    ws['A5'].fill = self.GATEWAY_FILL

                # Table headers
                current_row = 7
                headers = ["IP Address", "Device/VM Name", "Device Role", "Device Model", "Interface", "Type", "Status"]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col, value=header)
                    cell.font = self.HEADER_FONT
                    cell.fill = self.HEADER_FILL
                    cell.alignment = self.CENTER_ALIGN
                    cell.border = self.CELL_BORDER

                # Build IP list - either all IPs in range or just assigned
                prefix_network = netaddr.IPNetwork(str(prefix.prefix))
                prefix_size = prefix_network.prefixlen

                # Create lookup of assigned IPs by address
                assigned_ip_lookup = {}
                for ip in assigned_ips:
                    ip_addr_str = str(ip.address).split('/')[0]
                    assigned_ip_lookup[ip_addr_str] = ip

                # Determine if we should show all IPs (only for /24 or smaller to avoid huge lists)
                show_all_ips = show_unused_ips and prefix_size >= 24

                if show_all_ips:
                    # Generate all usable IPs in the prefix
                    all_ips_in_prefix = list(prefix_network.iter_hosts())
                else:
                    # Just use assigned IPs
                    all_ips_in_prefix = [netaddr.IPAddress(str(ip.address).split('/')[0]) for ip in assigned_ips]

                # IP address data rows
                current_row += 1
                data_start_row = current_row
                ip_count = 0

                for ip_addr in all_ips_in_prefix:
                    try:
                        ip_str = str(ip_addr)
                        assigned_ip = assigned_ip_lookup.get(ip_str)

                        if assigned_ip:
                            # This IP is assigned - show full details
                            device_info = self._get_ip_device_info(assigned_ip)
                            is_gateway = self._is_default_gateway(assigned_ip)

                            ip_display = str(assigned_ip.address)
                            if is_gateway:
                                ip_display = f"{assigned_ip.address} (GW)"

                            row_data = [
                                ip_display,
                                device_info['device_name'] or "Unassigned",
                                device_info['device_role'] or "N/A",
                                device_info['device_model'] or "N/A",
                                device_info['interface_name'] or "N/A",
                                device_info['device_type'] or "N/A",
                                device_info['status'] or "N/A"
                            ]

                            for col, value in enumerate(row_data, 1):
                                cell = ws.cell(row=current_row, column=col, value=value)
                                cell.border = self.CELL_BORDER

                                if is_gateway:
                                    cell.font = self.GATEWAY_FONT
                                    cell.fill = self.GATEWAY_FILL
                                else:
                                    cell.font = self.NORMAL_FONT
                                    if (current_row - data_start_row) % 2 == 1:
                                        cell.fill = self.ALT_ROW_FILL
                        else:
                            # This IP is available/unused - include prefix length
                            row_data = [
                                f"{ip_str}/{prefix_size}",
                                "",
                                "",
                                "",
                                "",
                                "",
                                "Available"
                            ]

                            for col, value in enumerate(row_data, 1):
                                cell = ws.cell(row=current_row, column=col, value=value)
                                cell.border = self.CELL_BORDER
                                cell.font = available_font
                                cell.fill = available_fill

                        current_row += 1
                        ip_count += 1

                    except Exception as e:
                        self.log_warning(f"Error processing IP {ip_addr}: {str(e)}")

                self.log_debug(f"Prefix {prefix.prefix}: {ip_count} IP addresses documented")
                sheets_created += 1

            except Exception as e:
                self.log_failure(f"Error creating sheet for prefix {prefix.prefix}: {str(e)}")

        self.log_info(f"Created {sheets_created} prefix sheets, skipped {sheets_skipped} empty prefixes")

    # ==========================================================================
    # Main Run Method
    # ==========================================================================

    def run(self, data, commit):
        """Main script execution."""

        # Check dependencies
        if not OPENPYXL_AVAILABLE:
            self.log_failure("openpyxl library is not installed. Please run: pip install openpyxl")
            return "ERROR: Missing required library 'openpyxl'"

        site = data['site']
        include_empty = data.get('include_empty_prefixes', True)
        show_unused_ips = data.get('show_unused_ips', True)
        branch = data.get('branch', None) if BRANCHING_AVAILABLE else None

        self.log_info(f"Starting network documentation generation for site: {site.name}")
        self.log_debug(f"Site ID: {site.id}, Slug: {site.slug}")

        if branch:
            self.log_info(f"Using branch: {branch.name}")
        else:
            self.log_info("Using main/production schema")

        try:
            # Initialize styles
            self._init_styles()

            # Define the main work function to run with or without branch context
            def do_work():
                # Collect data
                self.log_info("=" * 50)
                self.log_info("PHASE 1: Data Collection")
                self.log_info("=" * 50)

                prefixes = self._get_site_prefixes(site)
                vlans = self._get_site_vlans(site)
                orphan_vlans = self._get_orphan_vlans(site, prefixes)

                # Validate we have data to document
                if prefixes.count() == 0 and vlans.count() == 0:
                    self.log_failure(f"No network data found for site '{site.name}'")
                    return None, f"ERROR: No prefixes or VLANs found for site '{site.name}'"

                # Create workbook
                self.log_info("=" * 50)
                self.log_info("PHASE 2: Excel Document Generation")
                self.log_info("=" * 50)

                workbook = openpyxl.Workbook()

                # Pre-calculate sheet names for each prefix (needed for hyperlinks)
                prefix_sheet_names = {}
                for prefix in prefixes:
                    vlan_id = str(prefix.vlan.vid) if prefix.vlan else "NoVLAN"
                    description = prefix.description or str(prefix.prefix).replace('/', '_')
                    sheet_name = f"{vlan_id} - {description}"
                    for char in ['\\', '/', '*', '?', ':', '[', ']']:
                        sheet_name = sheet_name.replace(char, '-')
                    sheet_name = sheet_name[:31]
                    prefix_sheet_names[prefix.id] = sheet_name

                # Build sheets
                self._create_cover_page(workbook, site)
                self._create_summary_sheet(workbook, site, prefixes, orphan_vlans, prefix_sheet_names)
                self._create_prefix_sheets(workbook, prefixes, include_empty, prefix_sheet_names, show_unused_ips)

                return workbook, None

            # Execute with or without branch context
            if branch and BRANCHING_AVAILABLE:
                self.log_debug(f"Activating branch context: {branch.name}")
                with activate_branch(branch):
                    workbook, error = do_work()
            else:
                workbook, error = do_work()

            # Check for errors from data collection
            if error:
                return error

            # Save to buffer
            self.log_info("=" * 50)
            self.log_info("PHASE 3: File Generation")
            self.log_info("=" * 50)

            buffer = BytesIO()
            self.log_debug("Saving workbook to buffer...")
            workbook.save(buffer)
            self.log_debug("Workbook saved, seeking to start...")
            buffer.seek(0)
            file_content = buffer.getvalue()

            # Generate filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{site.slug}_network_documentation_{timestamp}.xlsx"

            self.log_info(f"Workbook saved to buffer, size: {len(file_content)} bytes")
            self.log_debug(f"Filename: {filename}")

            # Try to find the Job object and attach file for download
            from django.core.files.base import ContentFile
            from core.models import Job

            # Get current job from the request or find by current execution
            job = None

            # Try different ways to access the job
            if hasattr(self, 'job'):
                job = self.job
                self.log_debug("Found job via self.job")
            elif hasattr(self, 'request') and hasattr(self.request, 'job'):
                job = self.request.job
                self.log_debug("Found job via self.request.job")
            else:
                # Try to find the most recent job for this script
                try:
                    job = Job.objects.filter(
                        name=self.__class__.__name__,
                        status='running'
                    ).order_by('-created').first()
                    if job:
                        self.log_debug(f"Found job via query: {job.id}")
                except Exception as e:
                    self.log_debug(f"Could not find job via query: {e}")

            if job and hasattr(job, 'output_file'):
                # Save file to job's output_file field
                job.output_file.save(filename, ContentFile(file_content), save=True)
                self.log_success(f"Documentation ready for download: {filename} ({len(file_content):,} bytes)")
                return f"Documentation generated successfully!\n\nFile: {filename}\nSize: {len(file_content):,} bytes\n\nClick the download button above to save the file."

            # Fallback: Save to media directory with download URL
            self.log_debug("Job output_file not available, saving to media directory")
            import os
            from django.conf import settings

            media_root = getattr(settings, 'MEDIA_ROOT', '/opt/netbox/netbox/media')
            output_dir = os.path.join(media_root, 'script-outputs')
            os.makedirs(output_dir, exist_ok=True)

            file_path = os.path.join(output_dir, filename)
            with open(file_path, 'wb') as f:
                f.write(file_content)

            download_url = f"/media/script-outputs/{filename}"
            self.log_success(f"Documentation saved: {filename} ({len(file_content):,} bytes)")

            return f"Documentation generated successfully!\n\nFile: {filename}\nSize: {len(file_content):,} bytes\n\nDownload: {download_url}"

        except Exception as e:
            self.log_failure(f"Unexpected error during script execution: {str(e)}")
            self.log_debug(f"Exception type: {type(e).__name__}")

            # Log full traceback for debugging
            import traceback
            self.log_debug(f"Traceback:\n{traceback.format_exc()}")

            return f"ERROR: {str(e)}"


# Register the script
script = NetworkDocumentationScript
